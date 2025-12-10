"""
This script defines an ETL (Extract, Transform, Load) pipeline for processing
transaction data from an Excel file.

It includes data cleaning utilities, a Pydantic model for data validation,
and a memory-efficient generator to process large files row-by-row.
"""

import re
from datetime import date as type_date
from datetime import datetime
from typing import Any, Dict, Iterator, Optional, Self, Union

from openpyxl import load_workbook
from pydantic import BaseModel, Field, field_validator, model_validator

from SmartBudget.utils.logger import logger, setup_logging

# ---------- Utility Cleaners ----------


def clean_text(value: Optional[str]) -> Optional[str]:
    """
    Sanitizes a string by stripping leading/trailing whitespace.

    Args:
        value: The string to sanitize.

    Returns:
        The cleaned string, or None if the input is None or results in an
        empty string.
    """
    if value is None:
        return None
    value = str(value).strip()
    return value if value else None


def parse_quantity(value: Optional[Union[int, float, str]]) -> Optional[int]:
    """
    Parses a quantity value, handling various Excel formats.

    Excel can store numbers as strings ("3"), floats (3.0), or integers.
    This function robustly converts them to a standard integer format.

    Args:
        value: The value to parse, which could be an int, float, or string.

    Returns:
        The parsed integer quantity, or None if the input is None.

    Raises:
        ValueError: If the value cannot be converted to an integer.
    """
    if value is None:
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError) as exc:
        raise ValueError(f"Invalid quantity: '{value}'") from exc


def parse_price(value: Optional[Union[float, str]]) -> Optional[float]:
    """
    Parses a price value, removing currency symbols and other non-numeric characters.

    Handles formats like "$10.50", "10.5", or even "-5.00".

    Args:
        value: The price value to parse.

    Returns:
        The parsed float price, or None if the input is None.

    Raises:
        ValueError: If the value cannot be converted to a float after cleaning.
    """
    if value is None:
        return None
    try:
        # Remove any character that is not a digit, a dot, or a minus sign
        clean_value = re.sub(r"[^\d.\-]", "", str(value))
        return float(clean_value)
    except (ValueError, TypeError) as exc:
        raise ValueError(f"Invalid price: '{value}'") from exc


# ---------- Pydantic Data Model ----------


class TransactionRow(BaseModel):
    """
    Represents a single, validated transaction row from the Excel sheet.

    This Pydantic model defines the expected data structure and applies
    validation and cleaning rules to ensure data quality.
    """

    # --- Fields ---
    date: type_date = Field(
        ..., description="The date of the transaction.", alias="Date"
    )
    item: Optional[str] = Field(None, description="The item purchased.", alias="Item")
    category: Optional[str] = Field(
        None, description="The category of the expense.", alias="Category"
    )
    quantity: Optional[int] = Field(
        None, description="The quantity of the item.", alias="Quantity"
    )
    price: Optional[float] = Field(
        None, description="The price of the item.", alias="Price"
    )
    note: Optional[str] = Field(
        None, description="Additional notes about the transaction.", alias="Note"
    )
    payment_method: Optional[str] = Field(
        None, description="The method of payment.", alias="Payment Method"
    )

    # --- Configuration ---
    model_config = {
        "populate_by_name": True,
        "extra": "ignore",  # Drops "Unnamed: x" columns from Excel
    }

    # --- Field-level Validators ---

    @field_validator("item", "category", "note", "payment_method", mode="before")
    @classmethod
    def clean_strings(cls, v: str) -> Optional[str]:
        """Applies the clean_text utility to string-based fields."""
        return clean_text(v)

    @field_validator("quantity", mode="before")
    @classmethod
    def validate_quantity(cls, v: Optional[Union[int, float, str]]) -> Optional[int]:
        """Applies the parse_quantity utility to the quantity field."""
        return parse_quantity(v)

    @field_validator("price", mode="before")
    @classmethod
    def validate_price(cls, v: Optional[Union[float, str]]) -> Optional[float]:
        """Applies the parse_price utility to the price field."""
        return parse_price(v)

    @field_validator("date", mode="before")
    @classmethod
    def validate_date(cls, v: Any) -> type_date:
        """
        Validates and parses the date field from various formats.

        Handles pandas Timestamps, Excel's datetime strings, and ISO date strings.
        """
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, type_date):
            return v
        try:
            return type_date.fromisoformat(str(v))
        except (ValueError, TypeError) as exc:
            raise ValueError(f"Invalid date format: '{v}'") from exc

    # --- Row-level Validation ---

    @model_validator(mode="after")
    def check_business_rules(self) -> Self:
        """
        Enforces business rules on the validated data model.

        Raises:
            ValueError: If a business rule is violated.
        """
        if not self.item and not self.note:
            raise ValueError("Either 'Item' or 'Note' must be provided.")

        if self.quantity is not None and self.quantity < 0:
            raise ValueError("Quantity cannot be negative.")

        if self.price is not None and self.price < 0:
            raise ValueError("Price cannot be negative.")

        return self


# ---------- ETL Generator ----------


def process_transactions(
    filepath: str, sheet_name: str
) -> Iterator[Union[TransactionRow, Dict[str, Any]]]:
    """
    Reads an Excel file row-by-row, validates the data, and yields results.

    This generator function is highly memory-efficient as it avoids loading the
    entire file into memory. It uses `openpyxl` for direct, read-only access
    to the worksheet.

    Args:
        filepath: The path to the Excel workbook.
        sheet_name: The name of the sheet to process.

    Yields:
        An iterator that produces either a validated `TransactionRow` instance
        for a successful row or a dictionary containing error details for a
        failed row.

    Raises:
        ValueError: If the specified sheet is not found in the workbook.
    """
    try:
        wb = load_workbook(filename=filepath, read_only=True, data_only=True)
    except FileNotFoundError:
        logger.error(f"Error: The file '{filepath}' was not found.")
        return

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")

    ws = wb[sheet_name]

    # Assume the first row is the header and extract it.
    header_cells = ws[1]
    headers = [cell.value for cell in header_cells]

    # Iterate over data rows, skipping the header.
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_data = {headers[j]: cell.value for j, cell in enumerate(row)}

        try:
            # Validate the row data against the Pydantic model
            yield TransactionRow.model_validate(row_data)
        except Exception as e:
            # If validation fails, yield an error dictionary
            yield {"row_index": i, "error": str(e), "data": row_data}


# ---------- Main Execution ----------

if __name__ == "__main__":
    setup_logging()
    # --- Configuration ---
    FILEPATH = "data/Budget 2025-26.xlsx"
    SHEET_NAME = "Expense Tracker"

    # --- Processing ---
    logger.info(f"Starting ETL process for '{FILEPATH}' -> '{SHEET_NAME}'...")
    valid_rows = []
    error_rows = []

    # The generator processes the file row-by-row, minimizing memory use.
    # In a real-world ETL pipeline, you would typically stream these results
    # directly into a database or another destination instead of collecting
    # them in lists.
    for result in process_transactions(FILEPATH, SHEET_NAME):
        if isinstance(result, TransactionRow):
            valid_rows.append(result)
        else:
            error_rows.append(result)

    # --- Summary ---
    logger.info(f"\nProcessing complete.")
    logger.info(f"  - Valid rows: {len(valid_rows)}")
    logger.info(f"  - Error rows: {len(error_rows)}")

    # For debugging, you can inspect the errors.
    # Avoid printing the full lists for very large files.
    if error_rows:
        logger.error("\n--- Errors ---")
        for error in error_rows[:5]:  # Print first 5 errors
            logger.error(
                f"  Row {error['row_index']}: {error['error']} | Data: {error['data']}"
            )
